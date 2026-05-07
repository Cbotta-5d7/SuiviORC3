"""
============================================================================
SUIVI DE PRODUCTION ORC1 — Application de bureau
============================================================================
"""

import os
import sys
import json
import math
from datetime import datetime, date, time
from pathlib import Path
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ===========================================================================
# CONFIGURATION
# ===========================================================================

APP_TITLE   = "Suivi de production — Ligne ORC1"
APP_VERSION = "v1.1"
CONFIG_FILE = "config.json"
PASSWORD    = "4594"

COLORS = {
    "bleu":        "#2563EB",
    "bleu_fonce":  "#1E3A8A",
    "bleu_bg":     "#EFF6FF",
    "vert":        "#16A34A",
    "vert_bg":     "#DCFCE7",
    "rouge":       "#DC2626",
    "rouge_bg":    "#FEE2E2",
    "orange":      "#EA580C",
    "orange_bg":   "#FFEDD5",
    "violet":      "#7C3AED",
    "violet_bg":   "#EDE9FE",
    "gris_bg":     "#F1F5F9",
    "gris_panel":  "#F8FAFC",
    "gris_bord":   "#E2E8F0",
    "gris_label":  "#64748B",
    "gris_placeh": "#94A3B8",
    "gris_text":   "#1E293B",
    "blanc":       "#FFFFFF",
    "titre_bg":    "#0F172A",
}

LISTES = {
    "postes":       ["Matin", "Après-midi", "Nuit"],
    "pilotes":      ["Marie Lambert", "Thomas Bernard", "Julie Petit", "Karim Benali",
                     "Sophie Moreau", "Antoine Dubois", "Nadia Cherif", "Pierre Garnier"],
    "copilotes":    ["Lucas Martin", "Émilie Roux", "Mehdi Saïdi", "Camille Faure",
                     "Alexandre Vidal", "Fatima Khelifi", "Damien Robin", "Léa Fontaine"],
    "tailles":      ["40x60", "40x65", "50x70", "60x60", "65x65", "Traversin 140", "Autre…"],
    "fibres":       ["FIB-001", "FIB-002", "FIB-003", "FIB-MICRO"],
    "rattrapages":  ["Pochon / fibre", "Couture", "Emballage", "Presse à souder", "Presse à housse ZIP"],
    "equipements":  ["Chargeuse", "Carde", "Étaleur / Tour", "Coupe / Coupe circulaire",
                     "Tapis bascule / Tapis pesé N°1", "Enrouleur pochon", "Pesée / Tapis pesée N°2",
                     "Déviation pochon / Table de distribution", "Enfileuse pochon", "Kinna / Stroebel",
                     "Tapeuse", "Table rotative — Twin pack", "Enfileuse H100", "Enfileuse traversin",
                     "Presse ORS", "Presse à housse ZIP", "Cercleuse", "Enrouleuse traversin",
                     "OF taie", "Traçabilité fibre"],
    "types_panne":  ["Panne", "Maintenance hebdomadaire", "Maintenance préventive", "Réglage"],
    "intervenants": ["Pierre Garnier (technicien)", "Karim Benali (technicien)",
                     "Damien Robin (technicien)", "Service externe"],
    "nb_personnes": [str(i) for i in range(1, 13)],
}

EXCEL_HEADERS = [
    "ID", "Type", "Date", "Heure", "Ligne", "Pilote", "Co-pilote", "N° OF",
    "Poste", "Code produit", "Taille", "Code fibre", "Poids garn. (g)", "Durée OF (min)",
    "Réf. taie", "Nb pers.", "Heure début", "Heure fin", "Qté fab.", "Qté emb.",
    "2nd choix", "Manq. taies", "Manq. housse", "Déf. couture",
    "Nb rattrap.", "Durée rattrap. (min)", "Nb pb tech.", "Durée pb tech. (min)",
    "Nb pannes", "Durée pannes (min)",
    "Équipement", "Type panne", "Intervenant", "Détail panne",
    "Détail rattrapages", "Détail pb tech.", "Commentaire",
]


# ===========================================================================
# CONFIG
# ===========================================================================

def get_app_dir():
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def load_config():
    p = get_app_dir() / CONFIG_FILE
    if p.exists():
        try:
            with open(p, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(config):
    p = get_app_dir() / CONFIG_FILE
    try:
        with open(p, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Erreur sauvegarde config : {e}")


# ===========================================================================
# EXCEL STORE
# ===========================================================================

class ExcelStore:

    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self._ensure_file()

    def _ensure_file(self):
        if self.filepath.exists():
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "DATA"
        thin = Side(border_style='thin', color='1E293B')
        for i, h in enumerate(EXCEL_HEADERS, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1E3A8A')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(h) + 2)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        wb.save(self.filepath)

    def get_all(self):
        if not self.filepath.exists():
            return []
        try:
            wb = load_workbook(self.filepath, data_only=True)
        except Exception as e:
            messagebox.showerror("Erreur lecture", f"Impossible de lire le fichier Excel :\n{e}")
            return []
        ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = {}
            for i, h in enumerate(EXCEL_HEADERS):
                d[h] = row[i] if i < len(row) else None
            rows.append(d)
        wb.close()
        return rows

    def add(self, declaration):
        try:
            wb = load_workbook(self.filepath)
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier Excel :\n{e}")
            return False
        ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
        existing_ids = [int(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True)
                        if r[0] is not None and isinstance(r[0], (int, float))]
        declaration["ID"] = max(existing_ids) + 1 if existing_ids else 1
        next_row = ws.max_row + 1
        if ws.cell(row=2, column=1).value is None and ws.max_row == 2:
            next_row = 2
        for i, h in enumerate(EXCEL_HEADERS, start=1):
            value = declaration.get(h, "")
            cell = ws.cell(row=next_row, column=i, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if h == "Date" and isinstance(value, (datetime, date)):
                cell.number_format = 'dd/mm/yyyy'
            elif h in ("Heure", "Heure début", "Heure fin") and value:
                cell.number_format = 'hh:mm'
        try:
            wb.save(self.filepath)
            wb.close()
            return True
        except PermissionError:
            messagebox.showerror("Fichier verrouillé",
                                 "Le fichier Excel est ouvert dans Excel.\nFermez-le et réessayez.")
            return False
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'enregistrer :\n{e}")
            return False

    def update(self, row_id, declaration):
        try:
            wb = load_workbook(self.filepath)
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier Excel :\n{e}")
            return False
        ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
        target_row = None
        for idx in range(2, ws.max_row + 1):
            val = ws.cell(row=idx, column=1).value
            if val is not None and int(val) == int(row_id):
                target_row = idx
                break
        if target_row is None:
            messagebox.showerror("Erreur", f"Ligne ID={row_id} introuvable.")
            wb.close()
            return False
        for i, h in enumerate(EXCEL_HEADERS, start=1):
            value = declaration.get(h, "")
            cell = ws.cell(row=target_row, column=i, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if h == "Date" and isinstance(value, (datetime, date)):
                cell.number_format = 'dd/mm/yyyy'
            elif h in ("Heure", "Heure début", "Heure fin") and value:
                cell.number_format = 'hh:mm'
        try:
            wb.save(self.filepath)
            wb.close()
            return True
        except PermissionError:
            messagebox.showerror("Fichier verrouillé",
                                 "Le fichier Excel est ouvert dans Excel.\nFermez-le et réessayez.")
            return False
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'enregistrer :\n{e}")
            return False

    def delete(self, row_id):
        try:
            wb = load_workbook(self.filepath)
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier Excel :\n{e}")
            return False
        ws = wb["DATA"] if "DATA" in wb.sheetnames else wb.active
        target_row = None
        for idx in range(2, ws.max_row + 1):
            val = ws.cell(row=idx, column=1).value
            if val is not None and int(val) == int(row_id):
                target_row = idx
                break
        if target_row is None:
            messagebox.showerror("Erreur", f"Ligne ID={row_id} introuvable.")
            wb.close()
            return False
        ws.delete_rows(target_row)
        try:
            wb.save(self.filepath)
            wb.close()
            return True
        except PermissionError:
            messagebox.showerror("Fichier verrouillé",
                                 "Le fichier Excel est ouvert dans Excel.\nFermez-le et réessayez.")
            return False
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'enregistrer :\n{e}")
            return False


# ===========================================================================
# WIDGETS PERSONNALISÉS
# ===========================================================================

class GaugeCanvas(ctk.CTkCanvas):

    def __init__(self, parent, width=280, height=180, **kwargs):
        super().__init__(parent, width=width, height=height,
                         bg=COLORS["blanc"], highlightthickness=0, **kwargs)
        self.w = width
        self.h = height
        self._value = 0
        self.draw(0)

    def draw(self, value):
        self._value = value
        self.delete("all")
        cx, cy = self.w // 2, int(self.h * 0.88)
        radius = int(min(self.w, self.h * 2) * 0.38)
        thickness = 20
        self._draw_arc(cx, cy, radius, 180, 180 - 108, COLORS["rouge"], thickness)
        self._draw_arc(cx, cy, radius, 180 - 108, 180 - 135, COLORS["orange"], thickness)
        self._draw_arc(cx, cy, radius, 180 - 135, 0, COLORS["vert"], thickness)
        self.create_text(cx - radius - 8, cy + 10, text="0",
                         font=("Segoe UI", 8), fill=COLORS["gris_label"])
        self.create_text(cx, cy - radius - 14, text="75",
                         font=("Segoe UI", 8), fill=COLORS["gris_label"])
        self.create_text(cx + radius + 10, cy + 10, text="100",
                         font=("Segoe UI", 8), fill=COLORS["gris_label"])
        angle_deg = 180 - (value * 1.8)
        angle_rad = math.radians(angle_deg)
        needle_len = radius - thickness // 2 - 4
        x_end = cx + needle_len * math.cos(angle_rad)
        y_end = cy - needle_len * math.sin(angle_rad)
        self.create_line(cx, cy, x_end, y_end, width=3, fill=COLORS["gris_text"], capstyle="round")
        self.create_oval(cx - 7, cy - 7, cx + 7, cy + 7,
                         fill=COLORS["gris_text"], outline="")
        self.create_text(cx, cy - radius * 0.44, text=f"{int(value)}%",
                         font=("Segoe UI", 30, "bold"), fill=COLORS["gris_text"])
        self.create_text(cx, cy - radius * 0.44 + 26, text="TRS",
                         font=("Segoe UI", 9), fill=COLORS["gris_label"])

    def _draw_arc(self, cx, cy, radius, start_deg, extent_deg, color, thickness):
        x0, y0 = cx - radius, cy - radius
        x1, y1 = cx + radius, cy + radius
        extent = extent_deg - start_deg
        self.create_arc(x0, y0, x1, y1, start=start_deg, extent=extent,
                        style="arc", outline=color, width=thickness)


class StopBar(ctk.CTkFrame):

    def __init__(self, parent, icon, label, color, **kwargs):
        super().__init__(parent, fg_color="transparent", **kwargs)
        self.color = color

        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x")

        ctk.CTkLabel(top, text=icon, width=24, height=24,
                     font=("Segoe UI", 12),
                     fg_color=color, text_color=COLORS["blanc"],
                     corner_radius=5).pack(side="left")

        ctk.CTkLabel(top, text=label, font=("Segoe UI", 12, "bold"),
                     text_color=COLORS["gris_text"], anchor="w"
                     ).pack(side="left", padx=(8, 0), fill="x", expand=True)

        self.count_lbl = ctk.CTkLabel(top, text="0",
                                      font=("Segoe UI", 10),
                                      text_color=COLORS["gris_label"],
                                      fg_color=COLORS["gris_bg"], corner_radius=8,
                                      width=34, height=20)
        self.count_lbl.pack(side="left", padx=(0, 8))

        self.value_lbl = ctk.CTkLabel(top, text="0 min",
                                      font=("Segoe UI", 13, "bold"),
                                      text_color=COLORS["gris_text"], width=64)
        self.value_lbl.pack(side="right")

        self.bar = ctk.CTkProgressBar(self, height=6,
                                      fg_color=COLORS["gris_bg"],
                                      progress_color=color,
                                      corner_radius=3)
        self.bar.pack(fill="x", pady=(4, 0), padx=(32, 0))
        self.bar.set(0)

    def update(self, count, value, max_value):
        self.count_lbl.configure(text=f"{count}")
        self.value_lbl.configure(text=f"{value} min")
        self.bar.set(value / max_value if max_value > 0 else 0)


# ===========================================================================
# APPLICATION PRINCIPALE
# ===========================================================================

class App(ctk.CTk):

    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(APP_TITLE)
        self.geometry("1280x820")
        self.minsize(1100, 700)
        self.configure(fg_color=COLORS["gris_bg"])
        self.after(50, self._maximize)

        self.config_data = load_config()
        self.excel_path = self.config_data.get("excel_path")

        if not self.excel_path or not Path(self.excel_path).exists():
            self._select_excel_file()
            if not self.excel_path:
                self.destroy()
                return

        self.store = ExcelStore(self.excel_path)
        self._build_titlebar()
        self._build_main()
        self._refresh()

    def _maximize(self):
        try:
            self.state('zoomed')
        except Exception:
            try:
                self.attributes('-zoomed', True)
            except Exception:
                pass

    # -----------------------------------------------------------------------
    # FICHIER EXCEL
    # -----------------------------------------------------------------------

    def _select_excel_file(self):
        choice = messagebox.askyesnocancel(
            "Fichier de données",
            "Voulez-vous OUVRIR un fichier Excel existant ?\n\n"
            "• Oui → choisir un fichier existant\n"
            "• Non → créer un nouveau fichier\n"
            "• Annuler → quitter l'application"
        )
        if choice is None:
            return
        if choice:
            path = filedialog.askopenfilename(
                title="Sélectionner le fichier Excel",
                filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
            )
        else:
            path = filedialog.asksaveasfilename(
                title="Créer un nouveau fichier Excel",
                defaultextension=".xlsx",
                initialfile="Donnees_ORC1.xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx")]
            )
        if not path:
            return
        self.excel_path = path
        self.config_data["excel_path"] = path
        save_config(self.config_data)

    def _change_excel_file(self):
        if not self._ask_password("Changer de répertoire"):
            return
        old_path = self.excel_path
        self._select_excel_file()
        if self.excel_path and self.excel_path != old_path:
            self.store = ExcelStore(self.excel_path)
            self._refresh()
            self.path_label.configure(text=self._short_path(self.excel_path))

    def _short_path(self, p):
        p = str(p)
        return p if len(p) <= 55 else "…" + p[-53:]

    # -----------------------------------------------------------------------
    # MOT DE PASSE
    # -----------------------------------------------------------------------

    def _ask_password(self, title="Mot de passe requis"):
        dialog = ctk.CTkInputDialog(text="Mot de passe :", title=title)
        pwd = dialog.get_input()
        if pwd is None:
            return False
        if pwd != PASSWORD:
            messagebox.showerror("Accès refusé", "Mot de passe incorrect.")
            return False
        return True

    # -----------------------------------------------------------------------
    # BARRE DE TITRE
    # -----------------------------------------------------------------------

    def _build_titlebar(self):
        bar = ctk.CTkFrame(self, fg_color=COLORS["bleu_fonce"], corner_radius=0, height=52)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        left = ctk.CTkFrame(bar, fg_color="transparent")
        left.pack(side="left", padx=20, pady=8)

        ctk.CTkLabel(left, text="🏭", font=("Segoe UI", 20),
                     fg_color=COLORS["bleu"], corner_radius=8,
                     width=34, height=34).pack(side="left")

        ctk.CTkLabel(left, text=APP_TITLE,
                     font=("Segoe UI", 14, "bold"),
                     text_color=COLORS["blanc"]).pack(side="left", padx=(12, 0))

        ctk.CTkLabel(left, text=APP_VERSION,
                     font=("Segoe UI", 10),
                     text_color="#93C5FD").pack(side="left", padx=(10, 0))

        right = ctk.CTkFrame(bar, fg_color="transparent")
        right.pack(side="right", padx=20)

        self.path_label = ctk.CTkLabel(right, text=self._short_path(self.excel_path),
                                       font=("Segoe UI", 10),
                                       text_color="#BFDBFE")
        self.path_label.pack(side="left", padx=(0, 10))

        ctk.CTkButton(right, text="📁  Changer",
                      width=90, height=28,
                      font=("Segoe UI", 11),
                      fg_color=COLORS["bleu"], hover_color="#1D4ED8",
                      text_color=COLORS["blanc"], corner_radius=6,
                      command=self._change_excel_file).pack(side="left")

    # -----------------------------------------------------------------------
    # ZONE PRINCIPALE
    # -----------------------------------------------------------------------

    def _build_main(self):
        self.container = ctk.CTkFrame(self, fg_color=COLORS["gris_bg"], corner_radius=0)
        self.container.pack(fill="both", expand=True)
        self._build_dashboard()

    def _clear_container(self):
        for w in self.container.winfo_children():
            w.destroy()

    def _build_dashboard(self):
        self._clear_container()

        outer = ctk.CTkFrame(self.container, fg_color=COLORS["gris_bg"], corner_radius=0)
        outer.pack(fill="both", expand=True, padx=20, pady=12)

        # ── Ligne 0 : Hero compact ──────────────────────────────────────────
        hero = ctk.CTkFrame(outer, fg_color="transparent", height=38)
        hero.pack(fill="x", pady=(0, 8))
        hero.pack_propagate(False)

        ctk.CTkLabel(hero, text="Tableau de bord — ORC1",
                     font=("Segoe UI", 20, "bold"),
                     text_color=COLORS["gris_text"], anchor="w").pack(side="left")

        self.last_update_lbl = ctk.CTkLabel(hero, text="",
                                             font=("Segoe UI", 11),
                                             text_color=COLORS["gris_label"])
        self.last_update_lbl.pack(side="right")

        # ── Ligne 1 : KPI (hauteur fixe) ────────────────────────────────────
        kpi_outer = ctk.CTkFrame(outer, fg_color="transparent", height=210)
        kpi_outer.pack(fill="x", pady=(0, 8))
        kpi_outer.pack_propagate(False)
        kpi_outer.columnconfigure(0, weight=0)
        kpi_outer.columnconfigure(1, weight=1)
        kpi_outer.columnconfigure(2, weight=1)
        kpi_outer.rowconfigure(0, weight=1)

        # --- Colonne 0 : Jauge TRS ---
        gauge_card = ctk.CTkFrame(kpi_outer, fg_color=COLORS["blanc"],
                                  corner_radius=14, border_width=1,
                                  border_color=COLORS["gris_bord"])
        gauge_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.kpi_title_lbl = ctk.CTkLabel(gauge_card, text="TRS — Ligne ORC1",
                                           font=("Segoe UI", 12, "bold"),
                                           text_color=COLORS["gris_text"])
        self.kpi_title_lbl.pack(pady=(12, 0))

        self.gauge = GaugeCanvas(gauge_card, width=280, height=168)
        self.gauge.pack(padx=12)

        self.status_lbl = ctk.CTkLabel(gauge_card, text="—",
                                       font=("Segoe UI", 11, "bold"),
                                       fg_color=COLORS["vert_bg"],
                                       text_color=COLORS["vert"],
                                       corner_radius=999, width=130, height=26)
        self.status_lbl.pack(pady=(4, 10))

        # --- Colonne 1 : Quantités ---
        qty_card = ctk.CTkFrame(kpi_outer, fg_color=COLORS["bleu"],
                                corner_radius=14)
        qty_card.grid(row=0, column=1, sticky="nsew", padx=(0, 8))

        ctk.CTkLabel(qty_card, text="PIÈCES PRODUITES",
                     font=("Segoe UI", 11, "bold"),
                     text_color="#DBEAFE").pack(anchor="w", padx=18, pady=(16, 4))

        qty_row = ctk.CTkFrame(qty_card, fg_color="transparent")
        qty_row.pack(fill="x", padx=18)
        ctk.CTkLabel(qty_row, text="📦", font=("Segoe UI", 28),
                     text_color=COLORS["blanc"]).pack(side="left")
        self.qty_value_lbl = ctk.CTkLabel(qty_row, text="0",
                                          font=("Segoe UI", 42, "bold"),
                                          text_color=COLORS["blanc"])
        self.qty_value_lbl.pack(side="left", padx=(8, 6))
        ctk.CTkLabel(qty_row, text="pcs", font=("Segoe UI", 14),
                     text_color="#DBEAFE").pack(side="left", anchor="s", pady=(0, 8))

        self.qty_sub_lbl = ctk.CTkLabel(qty_card, text="—",
                                        font=("Segoe UI", 11),
                                        text_color="#DBEAFE")
        self.qty_sub_lbl.pack(anchor="w", padx=18, pady=(4, 0))

        ctk.CTkFrame(qty_card, fg_color="#3B82F6", height=1).pack(fill="x", padx=18, pady=10)

        ctk.CTkLabel(qty_card, text="ARRÊTS CUMULÉS",
                     font=("Segoe UI", 11, "bold"),
                     text_color="#DBEAFE").pack(anchor="w", padx=18, pady=(0, 4))

        arr_row = ctk.CTkFrame(qty_card, fg_color="transparent")
        arr_row.pack(fill="x", padx=18, pady=(0, 16))
        ctk.CTkLabel(arr_row, text="⏱", font=("Segoe UI", 22),
                     text_color=COLORS["blanc"]).pack(side="left")
        self.arr_total_lbl = ctk.CTkLabel(arr_row, text="0 min",
                                          font=("Segoe UI", 28, "bold"),
                                          text_color="#FED7AA")
        self.arr_total_lbl.pack(side="left", padx=(8, 6))

        # --- Colonne 2 : Détail arrêts ---
        stops_card = ctk.CTkFrame(kpi_outer, fg_color=COLORS["blanc"],
                                  corner_radius=14, border_width=1,
                                  border_color=COLORS["gris_bord"])
        stops_card.grid(row=0, column=2, sticky="nsew")

        ctk.CTkLabel(stops_card, text="DÉTAIL DES ARRÊTS",
                     font=("Segoe UI", 11, "bold"),
                     text_color=COLORS["gris_label"]).pack(anchor="w", padx=16, pady=(14, 8))

        ctk.CTkFrame(stops_card, fg_color=COLORS["gris_bord"], height=1).pack(fill="x", padx=16)

        bars = ctk.CTkFrame(stops_card, fg_color="transparent")
        bars.pack(fill="both", expand=True, padx=16, pady=8)

        self.bar_pannes = StopBar(bars, "⚡", "Pannes / Maintenance", COLORS["orange"])
        self.bar_pannes.pack(fill="x", pady=4)
        self.bar_ratt = StopBar(bars, "⏱", "Rattrapages", COLORS["violet"])
        self.bar_ratt.pack(fill="x", pady=4)
        self.bar_tech = StopBar(bars, "🔧", "Problèmes techniques", COLORS["rouge"])
        self.bar_tech.pack(fill="x", pady=4)

        # ── Ligne 2 : Boutons ────────────────────────────────────────────────
        btn_outer = ctk.CTkFrame(outer, fg_color="transparent", height=62)
        btn_outer.pack(fill="x", pady=(0, 8))
        btn_outer.pack_propagate(False)
        btn_outer.columnconfigure(0, weight=1)
        btn_outer.columnconfigure(1, weight=1)

        ctk.CTkButton(btn_outer,
                      text="📦   Déclarer une production",
                      font=("Segoe UI", 14, "bold"),
                      fg_color=COLORS["bleu"], hover_color=COLORS["bleu_fonce"],
                      text_color=COLORS["blanc"], corner_radius=12, height=56,
                      command=self._open_form_production
                      ).grid(row=0, column=0, sticky="ew", padx=(0, 6))

        ctk.CTkButton(btn_outer,
                      text="⚡   Déclarer une panne / maintenance",
                      font=("Segoe UI", 14, "bold"),
                      fg_color=COLORS["orange"], hover_color="#C2410C",
                      text_color=COLORS["blanc"], corner_radius=12, height=56,
                      command=self._open_form_panne
                      ).grid(row=0, column=1, sticky="ew", padx=(6, 0))

        # ── Ligne 3 : Historique (expand) ────────────────────────────────────
        hist_card = ctk.CTkFrame(outer, fg_color=COLORS["blanc"],
                                 corner_radius=14, border_width=1,
                                 border_color=COLORS["gris_bord"])
        hist_card.pack(fill="both", expand=True)

        hist_head = ctk.CTkFrame(hist_card, fg_color="transparent", height=44)
        hist_head.pack(fill="x", padx=16, pady=(10, 0))
        hist_head.pack_propagate(False)

        ctk.CTkLabel(hist_head, text="Déclarations récentes",
                     font=("Segoe UI", 13, "bold"),
                     text_color=COLORS["gris_text"], anchor="w").pack(side="left")

        self.filter_var = ctk.StringVar(value="all")
        filt_frame = ctk.CTkFrame(hist_head, fg_color="transparent")
        filt_frame.pack(side="right")

        self._filter_btns = {}
        for label, value in [("Tout", "all"), ("Production", "production"), ("Pannes", "panne")]:
            active = (value == "all")
            btn = ctk.CTkButton(filt_frame, text=label,
                                width=80, height=26, font=("Segoe UI", 11, "bold"),
                                corner_radius=6,
                                fg_color=COLORS["gris_text"] if active else COLORS["blanc"],
                                text_color=COLORS["blanc"] if active else COLORS["gris_label"],
                                hover_color=COLORS["gris_text"],
                                border_width=1, border_color=COLORS["gris_bord"],
                                command=lambda v=value: self._set_filter(v))
            btn.pack(side="left", padx=2)
            self._filter_btns[value] = btn

        ctk.CTkFrame(hist_card, fg_color=COLORS["gris_bord"], height=1).pack(fill="x", padx=0)

        COL_DEFS = self._table_cols()
        thead = ctk.CTkFrame(hist_card, fg_color=COLORS["gris_panel"], corner_radius=0, height=34)
        thead.pack(fill="x")
        thead.pack_propagate(False)
        for label, w, *_ in COL_DEFS:
            ctk.CTkLabel(thead, text=label, width=w,
                         font=("Segoe UI", 10, "bold"),
                         text_color=COLORS["gris_label"], anchor="w"
                         ).pack(side="left", padx=(10, 0))

        self.table_frame = ctk.CTkScrollableFrame(hist_card, fg_color=COLORS["blanc"],
                                                  corner_radius=0)
        self.table_frame.pack(fill="both", expand=True)

    @staticmethod
    def _table_cols():
        return [
            ("DATE / HEURE",   118),
            ("TYPE",            72),
            ("N° OF",          100),
            ("PILOTE",         130),
            ("DUR. PROD",       80),
            ("DUR. ARRÊTS",     84),
            ("QUANTITÉ",        80),
            ("ACTIONS",        110),
        ]

    def _set_filter(self, value):
        for v, btn in self._filter_btns.items():
            if v == value:
                btn.configure(fg_color=COLORS["gris_text"], text_color=COLORS["blanc"])
            else:
                btn.configure(fg_color=COLORS["blanc"], text_color=COLORS["gris_label"])
        self.filter_var.set(value)
        self._refresh_table()

    # -----------------------------------------------------------------------
    # RAFRAÎCHISSEMENT
    # -----------------------------------------------------------------------

    def _refresh(self):
        self._refresh_kpi()
        self._refresh_table()

    def _refresh_kpi(self):
        decls = self.store.get_all()
        prod_decls = [d for d in decls if d.get("Type") == "production"]

        total_qte = sum(int(d.get("Qté fab.", 0) or 0) for d in prod_decls
                        if isinstance(d.get("Qté fab.", 0), (int, float)))
        nb_of = len(prod_decls)
        self.qty_value_lbl.configure(text=f"{total_qte:,}".replace(",", " "))
        self.qty_sub_lbl.configure(text=f"{nb_of} OF terminé{'s' if nb_of > 1 else ''}")

        nb_p, dur_p = 0, 0
        nb_r, dur_r = 0, 0
        nb_t, dur_t = 0, 0
        for d in decls:
            try:
                nb_p  += int(d.get("Nb pannes")            or 0)
                dur_p += int(d.get("Durée pannes (min)")   or 0)
                nb_r  += int(d.get("Nb rattrap.")          or 0)
                dur_r += int(d.get("Durée rattrap. (min)") or 0)
                nb_t  += int(d.get("Nb pb tech.")          or 0)
                dur_t += int(d.get("Durée pb tech. (min)") or 0)
            except (ValueError, TypeError):
                pass

        total_arrets = dur_p + dur_r + dur_t
        max_v = max(dur_p, dur_r, dur_t, 1)
        self.bar_pannes.update(nb_p, dur_p, max_v)
        self.bar_ratt.update(nb_r, dur_r, max_v)
        self.bar_tech.update(nb_t, dur_t, max_v)
        self.arr_total_lbl.configure(text=f"{total_arrets} min")

        if total_arrets > 0 and nb_of > 0:
            trs = max(20, min(98, 98 - (total_arrets / 5)))
        else:
            trs = 88 if nb_of > 0 else 0

        self.gauge.draw(trs)

        if trs >= 75:
            self.status_lbl.configure(text="✓  Objectif atteint",
                                      fg_color=COLORS["vert_bg"], text_color=COLORS["vert"])
        elif trs >= 60:
            self.status_lbl.configure(text="⚠  En vigilance",
                                      fg_color=COLORS["orange_bg"], text_color=COLORS["orange"])
        else:
            self.status_lbl.configure(text="✗  Sous objectif",
                                      fg_color=COLORS["rouge_bg"], text_color=COLORS["rouge"])

        if decls:
            last = max(decls, key=lambda d: (str(d.get("Date") or ""), str(d.get("Heure") or "")))
            h = last.get("Heure")
            h_str = h.strftime("%H:%M") if isinstance(h, (time, datetime)) else str(h or "—")
            self.kpi_title_lbl.configure(text=f"TRS — Ligne ORC1  ·  dernière saisie {h_str}")
            dt = last.get("Date")
            dt_str = dt.strftime("%d/%m/%Y") if isinstance(dt, (datetime, date)) else str(dt or "")
            self.last_update_lbl.configure(text=f"Mise à jour : {dt_str} {h_str}")
        else:
            self.kpi_title_lbl.configure(text="TRS — Ligne ORC1")

    def _refresh_table(self):
        for w in self.table_frame.winfo_children():
            w.destroy()

        decls = self.store.get_all()
        flt = self.filter_var.get()
        if flt != "all":
            decls = [d for d in decls if d.get("Type") == flt]

        def sort_key(d):
            dt = d.get("Date")
            hr = d.get("Heure")
            d_s = dt.strftime("%Y%m%d") if isinstance(dt, (datetime, date)) else str(dt or "")
            h_s = hr.strftime("%H%M")   if isinstance(hr, (datetime, time))  else str(hr or "")
            return d_s + h_s

        decls = sorted(decls, key=sort_key, reverse=True)[:50]

        if not decls:
            ctk.CTkLabel(self.table_frame, text="Aucune déclaration",
                         font=("Segoe UI", 12, "italic"),
                         text_color=COLORS["gris_placeh"], height=50).pack(fill="x")
            return

        COL_DEFS = self._table_cols()

        for i, d in enumerate(decls):
            bg = COLORS["blanc"] if i % 2 == 0 else COLORS["gris_panel"]
            row = ctk.CTkFrame(self.table_frame, fg_color=bg, corner_radius=0, height=34)
            row.pack(fill="x")
            row.pack_propagate(False)

            dt = d.get("Date")
            hr = d.get("Heure")
            d_str = dt.strftime("%d/%m/%y") if isinstance(dt, (datetime, date)) else str(dt or "")
            h_str = hr.strftime("%H:%M")    if isinstance(hr, (datetime, time))  else str(hr or "")

            type_val   = d.get("Type") or ""
            type_label = "Production" if type_val == "production" else "Panne" if type_val == "panne" else "—"
            type_color = COLORS["bleu"]    if type_val == "production" else COLORS["orange"]
            type_bg    = COLORS["bleu_bg"] if type_val == "production" else COLORS["orange_bg"]

            qte = d.get("Qté fab.")
            qte_str = f"{int(qte):,} pcs".replace(",", " ") if isinstance(qte, (int, float)) and qte else "—"

            dur_prod   = int(d.get("Durée OF (min)") or 0)
            dur_arrets = (int(d.get("Durée pannes (min)") or 0)
                          + int(d.get("Durée rattrap. (min)") or 0)
                          + int(d.get("Durée pb tech. (min)") or 0))

            cells = [
                (f"{d_str} {h_str}", COL_DEFS[0][1], COLORS["gris_text"], None),
                (type_label,         COL_DEFS[1][1], type_color, type_bg),
                (str(d.get("N° OF") or "—"), COL_DEFS[2][1], COLORS["gris_text"], None),
                (str(d.get("Pilote") or "—"), COL_DEFS[3][1], COLORS["gris_text"], None),
                (f"{dur_prod} min" if dur_prod else "—", COL_DEFS[4][1], COLORS["bleu"], None),
                (f"{dur_arrets} min" if dur_arrets else "—", COL_DEFS[5][1],
                 COLORS["rouge"] if dur_arrets else COLORS["gris_placeh"], None),
                (qte_str, COL_DEFS[6][1], COLORS["gris_text"], None),
            ]

            for txt, w, color, bgcol in cells:
                if bgcol:
                    ctk.CTkLabel(row, text=txt, width=w,
                                 font=("Segoe UI", 10, "bold"),
                                 text_color=color, fg_color=bgcol,
                                 corner_radius=4, anchor="center"
                                 ).pack(side="left", padx=(10, 0))
                else:
                    ctk.CTkLabel(row, text=txt, width=w,
                                 font=("Segoe UI", 10),
                                 text_color=color, anchor="w"
                                 ).pack(side="left", padx=(10, 0))

            act_frame = ctk.CTkFrame(row, fg_color="transparent", width=COL_DEFS[7][1])
            act_frame.pack(side="left", padx=(10, 0))

            row_id = d.get("ID")
            ctk.CTkButton(act_frame, text="✏", width=34, height=24,
                          font=("Segoe UI", 12),
                          fg_color=COLORS["bleu_bg"], hover_color="#BFDBFE",
                          text_color=COLORS["bleu"], corner_radius=6,
                          command=lambda rid=row_id, dec=d: self._edit_row(rid, dec)
                          ).pack(side="left", padx=2)

            ctk.CTkButton(act_frame, text="🗑", width=34, height=24,
                          font=("Segoe UI", 12),
                          fg_color=COLORS["rouge_bg"], hover_color="#FECACA",
                          text_color=COLORS["rouge"], corner_radius=6,
                          command=lambda rid=row_id: self._delete_row(rid)
                          ).pack(side="left", padx=2)

            ctk.CTkFrame(self.table_frame, fg_color=COLORS["gris_bord"], height=1).pack(fill="x")

    def _edit_row(self, row_id, declaration):
        if not self._ask_password("Modifier une ligne"):
            return
        EditDialog(self, self.store, declaration, on_save=self._refresh)

    def _delete_row(self, row_id):
        if not self._ask_password("Supprimer une ligne"):
            return
        if not messagebox.askyesno("Confirmer",
                                   f"Supprimer définitivement la ligne ID={row_id} ?"):
            return
        if self.store.delete(row_id):
            self._refresh()

    def _open_form_production(self):
        FormProduction(self, self.store, on_save=self._refresh)

    def _open_form_panne(self):
        FormPanne(self, self.store, on_save=self._refresh)


# ===========================================================================
# DIALOGUE ÉDITION
# ===========================================================================

class EditDialog(ctk.CTkToplevel):

    def __init__(self, parent, store, declaration, on_save=None):
        super().__init__(parent)
        self.store = store
        self.declaration = dict(declaration)
        self.on_save = on_save
        self.row_id = declaration.get("ID")
        self.title(f"Modifier la déclaration #{self.row_id}")
        self.geometry("700x540")
        self.configure(fg_color=COLORS["gris_bg"])
        self.transient(parent)
        self.grab_set()
        self._build()

    def _build(self):
        is_prod = self.declaration.get("Type") == "production"
        color = COLORS["bleu"] if is_prod else COLORS["orange"]
        icon  = "📦" if is_prod else "⚡"
        ltype = "production" if is_prod else "panne / maintenance"

        head = ctk.CTkFrame(self, fg_color=color, corner_radius=0, height=60)
        head.pack(fill="x")
        head.pack_propagate(False)
        ctk.CTkLabel(head, text=f"{icon}   Modifier — {ltype}  ·  ID #{self.row_id}",
                     font=("Segoe UI", 15, "bold"),
                     text_color=COLORS["blanc"], anchor="w"
                     ).pack(side="left", padx=20)

        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["gris_bg"], corner_radius=0)
        scroll.pack(fill="both", expand=True)

        body = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                            corner_radius=12, border_width=1,
                            border_color=COLORS["gris_bord"])
        body.pack(fill="x", padx=20, pady=16)

        def field(parent, label, key, side="left", width=200):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.pack(side=side, padx=6, pady=4)
            ctk.CTkLabel(wrap, text=label.upper(),
                         font=("Segoe UI", 9, "bold"),
                         text_color=COLORS["gris_label"], anchor="w").pack(fill="x")
            e = ctk.CTkEntry(wrap, width=width, height=34,
                             font=("Segoe UI", 12),
                             fg_color=COLORS["blanc"],
                             border_color=COLORS["gris_bord"],
                             border_width=1, corner_radius=6,
                             text_color=COLORS["gris_text"])
            e.pack()
            val = self.declaration.get(key, "")
            if isinstance(val, (datetime, date)):
                val = val.strftime("%d/%m/%Y")
            elif isinstance(val, time):
                val = val.strftime("%H:%M")
            e.insert(0, str(val) if val else "")
            return e, key

        r1 = ctk.CTkFrame(body, fg_color="transparent")
        r1.pack(fill="x", padx=16, pady=(16, 0))
        self.f_date,   _ = field(r1, "Date",   "Date",   width=150)
        self.f_pilote, _ = field(r1, "Pilote", "Pilote", width=200)
        self.f_poste,  _ = field(r1, "Poste",  "Poste",  width=150)

        if is_prod:
            r2 = ctk.CTkFrame(body, fg_color="transparent")
            r2.pack(fill="x", padx=16)
            self.f_numof,  _ = field(r2, "N° OF",         "N° OF",          width=150)
            self.f_qtefab, _ = field(r2, "Qté fabriquée", "Qté fab.",        width=150)
            self.f_qteemb, _ = field(r2, "Qté emballée",  "Qté emb.",        width=150)
            r3 = ctk.CTkFrame(body, fg_color="transparent")
            r3.pack(fill="x", padx=16)
            self.f_hdebut, _ = field(r3, "Heure début",   "Heure début",    width=150)
            self.f_hfin,   _ = field(r3, "Heure fin",     "Heure fin",      width=150)
            self.f_duree,  _ = field(r3, "Durée OF (min)", "Durée OF (min)", width=150)
        else:
            r2 = ctk.CTkFrame(body, fg_color="transparent")
            r2.pack(fill="x", padx=16)
            self.f_equip,  _ = field(r2, "Équipement",    "Équipement",         width=200)
            self.f_type_p, _ = field(r2, "Type panne",    "Type panne",         width=200)
            r3 = ctk.CTkFrame(body, fg_color="transparent")
            r3.pack(fill="x", padx=16)
            self.f_hdebut,  _ = field(r3, "Heure début",        "Heure début",        width=150)
            self.f_hfin,    _ = field(r3, "Heure fin",          "Heure fin",          width=150)
            self.f_duree_p, _ = field(r3, "Durée pannes (min)", "Durée pannes (min)", width=150)
            r4 = ctk.CTkFrame(body, fg_color="transparent")
            r4.pack(fill="x", padx=16)
            self.f_detail, _ = field(r4, "Détail panne", "Détail panne", width=500, side="left")

        ctk.CTkFrame(body, fg_color=COLORS["gris_bord"], height=1).pack(fill="x", padx=16, pady=(12, 0))
        ctk.CTkLabel(body, text="COMMENTAIRE",
                     font=("Segoe UI", 9, "bold"),
                     text_color=COLORS["gris_label"], anchor="w"
                     ).pack(anchor="w", padx=22, pady=(8, 2))
        self.f_comment = ctk.CTkTextbox(body, height=70,
                                        font=("Segoe UI", 11),
                                        fg_color=COLORS["blanc"],
                                        border_color=COLORS["gris_bord"],
                                        border_width=1, corner_radius=6)
        self.f_comment.pack(fill="x", padx=20, pady=(0, 16))
        existing = self.declaration.get("Commentaire", "")
        if existing:
            self.f_comment.insert("1.0", str(existing))

        footer = ctk.CTkFrame(self, fg_color=COLORS["gris_panel"], corner_radius=0, height=64)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        ctk.CTkButton(footer, text="✓   Enregistrer les modifications",
                      font=("Segoe UI", 12, "bold"),
                      fg_color=color,
                      hover_color=COLORS["bleu_fonce"] if is_prod else "#C2410C",
                      text_color=COLORS["blanc"],
                      corner_radius=8, height=40, width=220,
                      command=self._save).pack(side="right", padx=16, pady=12)

        ctk.CTkButton(footer, text="Annuler",
                      font=("Segoe UI", 12),
                      fg_color=COLORS["blanc"], hover_color=COLORS["gris_bg"],
                      text_color=COLORS["gris_label"],
                      border_width=1, border_color=COLORS["gris_bord"],
                      corner_radius=8, height=40, width=100,
                      command=self.destroy).pack(side="right", pady=12)

    def _save(self):
        is_prod = self.declaration.get("Type") == "production"
        decl = dict(self.declaration)

        date_str = self.f_date.get().strip()
        try:
            decl["Date"] = datetime.strptime(date_str, "%d/%m/%Y").date()
        except ValueError:
            messagebox.showerror("Date invalide", "Format attendu : jj/mm/aaaa")
            return

        decl["Pilote"] = self.f_pilote.get().strip()
        decl["Poste"]  = self.f_poste.get().strip()

        def to_int(s):
            try:
                return int(float(s.strip())) if s and s.strip() else 0
            except Exception:
                return 0

        def parse_heure(s):
            s = (s or "").strip().replace("h", ":")
            if ":" not in s:
                return None
            try:
                p = s.split(":")
                return time(int(p[0]), int(p[1]))
            except Exception:
                return None

        if is_prod:
            decl["N° OF"]          = self.f_numof.get().strip()
            decl["Qté fab."]       = to_int(self.f_qtefab.get())
            decl["Qté emb."]       = to_int(self.f_qteemb.get())
            decl["Durée OF (min)"] = to_int(self.f_duree.get())
            hd = parse_heure(self.f_hdebut.get())
            hf = parse_heure(self.f_hfin.get())
            if hd: decl["Heure début"] = hd
            if hf: decl["Heure fin"]   = hf
        else:
            decl["Équipement"]         = self.f_equip.get().strip()
            decl["Type panne"]         = self.f_type_p.get().strip()
            decl["Durée pannes (min)"] = to_int(self.f_duree_p.get())
            decl["Détail panne"]       = self.f_detail.get().strip()
            hd = parse_heure(self.f_hdebut.get())
            hf = parse_heure(self.f_hfin.get())
            if hd: decl["Heure début"] = hd
            if hf: decl["Heure fin"]   = hf

        decl["Commentaire"] = self.f_comment.get("1.0", "end").strip()

        if self.store.update(self.row_id, decl):
            messagebox.showinfo("Enregistré", "Modification enregistrée.")
            if self.on_save:
                self.on_save()
            self.destroy()


# ===========================================================================
# HELPERS FORMULAIRES
# ===========================================================================

def _form_section(parent, num, icon, title, color=None, optional=False, hint=""):
    sec = ctk.CTkFrame(parent, fg_color=color or COLORS["bleu_bg"],
                       corner_radius=8, height=40)
    sec.pack(fill="x", padx=0, pady=(0, 2))
    sec.pack_propagate(False)
    ctk.CTkLabel(sec, text=f"  {num}  ", width=32,
                 font=("Segoe UI", 12, "bold"),
                 fg_color=color or COLORS["bleu"],
                 text_color=COLORS["blanc"],
                 corner_radius=6).pack(side="left", padx=(10, 0))
    ctk.CTkLabel(sec, text=f"{icon}  {title}",
                 font=("Segoe UI", 13, "bold"),
                 text_color=color or COLORS["bleu_fonce"]).pack(side="left", padx=10)
    if hint:
        ctk.CTkLabel(sec, text=hint,
                     font=("Segoe UI", 10),
                     text_color=COLORS["gris_placeh"]).pack(side="left")


def _form_field(parent, label_text, side="left", width=200, expand=False, default=""):
    wrap = ctk.CTkFrame(parent, fg_color="transparent")
    wrap.pack(side=side, padx=5, pady=4, fill="x" if expand else None, expand=expand)
    ctk.CTkLabel(wrap, text=label_text.upper(),
                 font=("Segoe UI", 9, "bold"),
                 text_color=COLORS["rouge"] if "*" in label_text else COLORS["gris_label"],
                 anchor="w").pack(fill="x", pady=(0, 3))
    e = ctk.CTkEntry(wrap, width=width, height=34,
                     font=("Segoe UI", 12),
                     fg_color=COLORS["blanc"],
                     border_color=COLORS["gris_bord"],
                     border_width=1, corner_radius=6,
                     text_color=COLORS["gris_text"])
    e.pack(fill="x")
    if default:
        e.insert(0, default)
    return e


def _form_dropdown(parent, label_text, values, side="left", width=200,
                   expand=False, default=""):
    wrap = ctk.CTkFrame(parent, fg_color="transparent")
    wrap.pack(side=side, padx=5, pady=4, fill="x" if expand else None, expand=expand)
    ctk.CTkLabel(wrap, text=label_text.upper(),
                 font=("Segoe UI", 9, "bold"),
                 text_color=COLORS["rouge"] if "*" in label_text else COLORS["gris_label"],
                 anchor="w").pack(fill="x", pady=(0, 3))
    var = ctk.StringVar(value=default if default else "Sélectionner…")
    combo = ctk.CTkComboBox(wrap, width=width, height=34,
                            values=values, variable=var,
                            font=("Segoe UI", 12),
                            fg_color=COLORS["blanc"],
                            border_color=COLORS["gris_bord"],
                            button_color=COLORS["bleu"],
                            button_hover_color=COLORS["bleu_fonce"],
                            dropdown_fg_color=COLORS["blanc"],
                            dropdown_text_color=COLORS["gris_text"],
                            text_color=COLORS["gris_text"],
                            state="readonly")
    combo.pack(fill="x")
    return combo


def _form_auto(parent, label_text, value, side="left", width=180):
    wrap = ctk.CTkFrame(parent, fg_color="transparent")
    wrap.pack(side=side, padx=5, pady=4)
    ctk.CTkLabel(wrap, text=label_text.upper(),
                 font=("Segoe UI", 9, "bold"),
                 text_color=COLORS["gris_label"], anchor="w").pack(fill="x", pady=(0, 3))
    lbl = ctk.CTkLabel(wrap, text=f"{value}  · auto",
                       width=width, height=34,
                       font=("Segoe UI", 12, "bold"),
                       text_color=COLORS["bleu_fonce"],
                       fg_color=COLORS["bleu_bg"], corner_radius=6,
                       anchor="w")
    lbl.pack(fill="x")
    lbl._auto_value = value
    return lbl


# ===========================================================================
# FORMULAIRE PRODUCTION
# ===========================================================================

class FormProduction(ctk.CTkToplevel):

    def __init__(self, parent, store, on_save=None):
        super().__init__(parent)
        self.store   = store
        self.on_save = on_save
        self.rattrap_rows  = []
        self.probleme_rows = []
        self.title("Déclarer une production — ORC1")
        self.geometry("1200x850")
        self.configure(fg_color=COLORS["gris_bg"])
        self.transient(parent)
        self.grab_set()
        self.after(60, self._maximize)
        self._build()

    def _maximize(self):
        try:
            self.state('zoomed')
        except Exception:
            try:
                self.attributes('-zoomed', True)
            except Exception:
                pass

    def _build(self):
        head = ctk.CTkFrame(self, fg_color=COLORS["bleu_fonce"], corner_radius=0, height=60)
        head.pack(fill="x")
        head.pack_propagate(False)
        ctk.CTkLabel(head, text="📦   Nouvelle déclaration de production — ORC1",
                     font=("Segoe UI", 16, "bold"),
                     text_color=COLORS["blanc"], anchor="w"
                     ).pack(side="left", padx=24)

        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["gris_bg"], corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=20, pady=12)

        _form_section(scroll, "1", "👤", "Identification", color=COLORS["bleu"])
        s1 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s1.pack(fill="x", pady=(0, 10))
        r = ctk.CTkFrame(s1, fg_color="transparent")
        r.pack(fill="x", padx=16, pady=10)
        self.f_date     = _form_field(r, "Date *",   width=160, default=date.today().strftime("%d/%m/%Y"))
        self.f_poste    = _form_dropdown(r, "Poste *",  LISTES["postes"],  width=160, default="Matin")
        self.f_pilote   = _form_dropdown(r, "Pilote *", LISTES["pilotes"], width=220)
        self.f_copilote = _form_dropdown(r, "Co-pilote", LISTES["copilotes"], width=220)

        _form_section(scroll, "2", "📦", "Ordre de fabrication", color=COLORS["bleu"])
        s2 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s2.pack(fill="x", pady=(0, 10))

        r2a = ctk.CTkFrame(s2, fg_color="transparent")
        r2a.pack(fill="x", padx=16, pady=(10, 0))
        self.f_numof    = _form_field(r2a, "N° OF *",        width=160)
        self.f_codeprod = _form_field(r2a, "Code produit",   width=160)
        self.f_taille   = _form_dropdown(r2a, "Taille",      LISTES["tailles"], width=160)
        self.f_fibre    = _form_dropdown(r2a, "Code fibre",  LISTES["fibres"],  width=160)
        self.f_poids    = _form_field(r2a, "Poids garn. (g)", width=140)

        r2b = ctk.CTkFrame(s2, fg_color="transparent")
        r2b.pack(fill="x", padx=16, pady=(0, 0))
        self.f_reftaie   = _form_field(r2b, "Réf. taie",   width=160)
        self.f_nbpers    = _form_dropdown(r2b, "Nb personnes", LISTES["nb_personnes"], width=120)
        self.f_hdebut    = _form_field(r2b, "Heure début", width=130)
        self.f_hfin      = _form_field(r2b, "Heure fin",   width=130)
        self.f_duree_lbl = _form_auto(r2b, "Durée OF",     "—", width=130)

        r2c = ctk.CTkFrame(s2, fg_color="transparent")
        r2c.pack(fill="x", padx=16, pady=(0, 10))
        self.f_qtefab       = _form_field(r2c, "Qté fabriquée",  width=160)
        self.f_qteemb       = _form_field(r2c, "Qté emballée",   width=160)
        self.f_equiv_lbl    = _form_auto(r2c, "Équivalence",     "—", width=130)
        self.f_2nd          = _form_field(r2c, "2nd choix",      width=120)
        self.f_cad_min_lbl  = _form_auto(r2c, "Cadence /min",    "—", width=120)
        self.f_cad_pers_lbl = _form_auto(r2c, "Cadence /pers/h", "—", width=130)

        self.f_hdebut.bind("<KeyRelease>", self._recompute)
        self.f_hfin.bind("<KeyRelease>",   self._recompute)
        self.f_qtefab.bind("<KeyRelease>",  self._recompute)
        self.f_qteemb.bind("<KeyRelease>",  self._recompute)
        self.f_nbpers.configure(command=lambda v: self._recompute())

        _form_section(scroll, "3", "🧵", "Matière première", color=COLORS["violet"], optional=True)
        s3 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s3.pack(fill="x", pady=(0, 10))
        r3 = ctk.CTkFrame(s3, fg_color="transparent")
        r3.pack(fill="x", padx=16, pady=10)
        self.f_manq_taies  = _form_field(r3, "Manquants taies",  width=200)
        self.f_manq_housse = _form_field(r3, "Manquants housse", width=200)
        self.f_def_couture = _form_field(r3, "Défauts couture",  width=200)

        _form_section(scroll, "4", "⏱", "Rattrapages", color=COLORS["violet"], optional=True)
        s4 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s4.pack(fill="x", pady=(0, 10))
        self.rattrap_container = ctk.CTkFrame(s4, fg_color="transparent")
        self.rattrap_container.pack(fill="x", padx=16, pady=(8, 0))
        ctk.CTkButton(s4, text="＋  Ajouter un rattrapage",
                      font=("Segoe UI", 11, "bold"),
                      fg_color=COLORS["violet_bg"], hover_color="#DDD6FE",
                      text_color=COLORS["violet"],
                      border_width=1, border_color="#C4B5FD",
                      corner_radius=6, height=32,
                      command=self._add_rattrap).pack(fill="x", padx=16, pady=(4, 10))
        self._add_rattrap()

        _form_section(scroll, "5", "🔧", "Problèmes techniques", color=COLORS["rouge"], optional=True)
        s5 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s5.pack(fill="x", pady=(0, 10))
        self.probleme_container = ctk.CTkFrame(s5, fg_color="transparent")
        self.probleme_container.pack(fill="x", padx=16, pady=(8, 0))
        ctk.CTkButton(s5, text="＋  Ajouter un problème technique",
                      font=("Segoe UI", 11, "bold"),
                      fg_color=COLORS["rouge_bg"], hover_color="#FECACA",
                      text_color=COLORS["rouge"],
                      border_width=1, border_color="#FCA5A5",
                      corner_radius=6, height=32,
                      command=self._add_probleme).pack(fill="x", padx=16, pady=(4, 10))
        self._add_probleme()

        _form_section(scroll, "6", "💬", "Commentaire général", color=COLORS["gris_label"], optional=True)
        s6 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s6.pack(fill="x", pady=(0, 10))
        self.f_comment = ctk.CTkTextbox(s6, height=80,
                                        font=("Segoe UI", 12),
                                        fg_color=COLORS["blanc"],
                                        border_color=COLORS["gris_bord"],
                                        border_width=1, corner_radius=6)
        self.f_comment.pack(fill="x", padx=16, pady=10)

        footer = ctk.CTkFrame(self, fg_color=COLORS["blanc"],
                              corner_radius=0, height=68,
                              border_width=1, border_color=COLORS["gris_bord"])
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        ctk.CTkLabel(footer, text="  * Champs obligatoires",
                     font=("Segoe UI", 11),
                     text_color=COLORS["gris_label"]).pack(side="left", padx=20)

        ctk.CTkButton(footer, text="Annuler",
                      font=("Segoe UI", 12),
                      fg_color=COLORS["gris_bg"], hover_color=COLORS["gris_bord"],
                      text_color=COLORS["gris_label"],
                      border_width=1, border_color=COLORS["gris_bord"],
                      corner_radius=8, height=42, width=110,
                      command=self.destroy).pack(side="right", padx=(0, 16), pady=13)

        ctk.CTkButton(footer, text="✓   Enregistrer la production",
                      font=("Segoe UI", 13, "bold"),
                      fg_color=COLORS["bleu"], hover_color=COLORS["bleu_fonce"],
                      text_color=COLORS["blanc"],
                      corner_radius=8, height=42, width=240,
                      command=self._save).pack(side="right", padx=(0, 8), pady=13)

    def _add_rattrap(self):
        if len(self.rattrap_rows) >= 5:
            messagebox.showinfo("Limite", "Maximum 5 rattrapages.")
            return
        self._add_repeater(self.rattrap_container, self.rattrap_rows, LISTES["rattrapages"])

    def _add_probleme(self):
        if len(self.probleme_rows) >= 10:
            messagebox.showinfo("Limite", "Maximum 10 problèmes.")
            return
        self._add_repeater(self.probleme_container, self.probleme_rows, LISTES["equipements"])

    def _add_repeater(self, container, store_list, options):
        idx = len(store_list) + 1
        row = ctk.CTkFrame(container, fg_color=COLORS["gris_panel"],
                           corner_radius=8, border_width=1, border_color=COLORS["gris_bord"])
        row.pack(fill="x", pady=3)

        num = ctk.CTkLabel(row, text=str(idx), width=22,
                           font=("Segoe UI", 11, "bold"),
                           text_color=COLORS["gris_label"])
        num.pack(side="left", padx=(10, 6), pady=8)

        var = ctk.StringVar(value="Sélectionner…")
        combo = ctk.CTkComboBox(row, width=240, height=30, values=options,
                                variable=var, font=("Segoe UI", 11),
                                fg_color=COLORS["blanc"],
                                border_color=COLORS["gris_bord"],
                                button_color=COLORS["bleu"], state="readonly")
        combo.pack(side="left", padx=4, pady=8)

        comment = ctk.CTkEntry(row, height=30, font=("Segoe UI", 11),
                               placeholder_text="Commentaire (optionnel)",
                               fg_color=COLORS["blanc"],
                               border_color=COLORS["gris_bord"])
        comment.pack(side="left", padx=4, pady=8, fill="x", expand=True)

        duree = ctk.CTkEntry(row, width=72, height=30, font=("Segoe UI", 11),
                             placeholder_text="min",
                             fg_color=COLORS["blanc"],
                             border_color=COLORS["gris_bord"])
        duree.pack(side="left", padx=4, pady=8)

        def remove():
            row.destroy()
            store_list[:] = [r for r in store_list if r["frame"] != row]
            for i, r in enumerate(store_list, 1):
                r["num_lbl"].configure(text=str(i))

        ctk.CTkButton(row, text="✕", width=30, height=30,
                      fg_color="transparent", hover_color=COLORS["rouge_bg"],
                      text_color=COLORS["gris_placeh"],
                      command=remove).pack(side="left", padx=(4, 10), pady=8)

        store_list.append({"frame": row, "num_lbl": num,
                           "type": combo, "comment": comment, "duree": duree})

    def _recompute(self, *_):
        try:
            hd = self._parse_time(self.f_hdebut.get())
            hf = self._parse_time(self.f_hfin.get())
            if hd and hf:
                duree = (hf[0] * 60 + hf[1]) - (hd[0] * 60 + hd[1])
                if duree < 0:
                    duree += 1440
                self.f_duree_lbl.configure(text=f"{duree} min  · auto")
                self.f_duree_lbl._auto_value = duree
            else:
                self.f_duree_lbl.configure(text="—  · auto")
                self.f_duree_lbl._auto_value = 0
        except Exception:
            pass

        try:
            qte   = float(self.f_qtefab.get() or 0)
            duree = self.f_duree_lbl._auto_value or 0
            nb_p  = float(self.f_nbpers.get() or 0)
            if qte and duree:
                self.f_cad_min_lbl.configure(text=f"{qte/duree:.2f}  · auto")
            else:
                self.f_cad_min_lbl.configure(text="—  · auto")
            if qte and duree and nb_p:
                self.f_cad_pers_lbl.configure(text=f"{int(qte/nb_p/(duree/60))}  · auto")
            else:
                self.f_cad_pers_lbl.configure(text="—  · auto")
        except Exception:
            pass

        try:
            qte     = float(self.f_qtefab.get() or 0)
            qte_emb = float(self.f_qteemb.get() or 0)
            if qte and qte_emb:
                self.f_equiv_lbl.configure(text=f"{(qte_emb/qte)*100:.1f}%  · auto")
            else:
                self.f_equiv_lbl.configure(text="—  · auto")
        except Exception:
            pass

    def _parse_time(self, s):
        s = (s or "").strip().replace("h", ":").replace(".", ":")
        if ":" not in s:
            return None
        try:
            p = s.split(":")
            return (int(p[0]), int(p[1]))
        except Exception:
            return None

    def _to_int(self, s):
        try:
            return int(float(s.strip())) if s and s.strip() else 0
        except Exception:
            return 0

    def _save(self):
        numof  = self.f_numof.get().strip()
        pilote = self.f_pilote.get()
        if not numof:
            messagebox.showerror("Champ obligatoire", "Le N° d'OF est obligatoire.")
            return
        if pilote == "Sélectionner…":
            messagebox.showerror("Champ obligatoire", "Le pilote est obligatoire.")
            return

        date_str = self.f_date.get().strip()
        try:
            d_obj = datetime.strptime(date_str, "%d/%m/%Y").date()
        except ValueError:
            messagebox.showerror("Date invalide", "Format attendu : jj/mm/aaaa")
            return

        decl = {h: "" for h in EXCEL_HEADERS}
        decl["Type"]      = "production"
        decl["Date"]      = d_obj
        decl["Heure"]     = datetime.now().time().replace(microsecond=0)
        decl["Ligne"]     = "ORC1"
        decl["Pilote"]    = pilote
        decl["Co-pilote"] = self.f_copilote.get() if self.f_copilote.get() != "Sélectionner…" else ""
        decl["N° OF"]     = numof
        decl["Poste"]     = self.f_poste.get()
        decl["Code produit"]    = self.f_codeprod.get()
        decl["Taille"]          = self.f_taille.get() if self.f_taille.get() != "Sélectionner…" else ""
        decl["Code fibre"]      = self.f_fibre.get()  if self.f_fibre.get()  != "Sélectionner…" else ""
        decl["Poids garn. (g)"] = self._to_int(self.f_poids.get())
        decl["Réf. taie"]       = self.f_reftaie.get()
        decl["Nb pers."]        = self._to_int(self.f_nbpers.get())

        hd = self._parse_time(self.f_hdebut.get())
        hf = self._parse_time(self.f_hfin.get())
        if hd: decl["Heure début"] = time(hd[0], hd[1])
        if hf: decl["Heure fin"]   = time(hf[0], hf[1])

        decl["Durée OF (min)"] = self.f_duree_lbl._auto_value or 0
        decl["Qté fab."]  = self._to_int(self.f_qtefab.get())
        decl["Qté emb."]  = self._to_int(self.f_qteemb.get())
        decl["2nd choix"] = self._to_int(self.f_2nd.get())
        decl["Manq. taies"]  = self._to_int(self.f_manq_taies.get())
        decl["Manq. housse"] = self._to_int(self.f_manq_housse.get())
        decl["Déf. couture"] = self._to_int(self.f_def_couture.get())

        nb_r, dur_r, det_r = 0, 0, []
        for r in self.rattrap_rows:
            t = r["type"].get()
            if t and t != "Sélectionner…":
                nb_r += 1
                d = self._to_int(r["duree"].get())
                dur_r += d
                det_r.append(f"{t} ({d}min)" + (f" – {r['comment'].get()}" if r["comment"].get() else ""))
        decl["Nb rattrap."]          = nb_r
        decl["Durée rattrap. (min)"] = dur_r
        decl["Détail rattrapages"]   = "; ".join(det_r)

        nb_p, dur_p, det_p = 0, 0, []
        for r in self.probleme_rows:
            t = r["type"].get()
            if t and t != "Sélectionner…":
                nb_p += 1
                d = self._to_int(r["duree"].get())
                dur_p += d
                det_p.append(f"{t} ({d}min)" + (f" – {r['comment'].get()}" if r["comment"].get() else ""))
        decl["Nb pb tech."]          = nb_p
        decl["Durée pb tech. (min)"] = dur_p
        decl["Détail pb tech."]      = "; ".join(det_p)

        decl["Commentaire"] = self.f_comment.get("1.0", "end").strip()

        if self.store.add(decl):
            messagebox.showinfo("Enregistré", "Déclaration de production enregistrée.")
            if self.on_save:
                self.on_save()
            self.destroy()


# ===========================================================================
# FORMULAIRE PANNE
# ===========================================================================

class FormPanne(ctk.CTkToplevel):

    def __init__(self, parent, store, on_save=None):
        super().__init__(parent)
        self.store   = store
        self.on_save = on_save
        self.title("Déclarer une panne / maintenance — ORC1")
        self.geometry("1000x700")
        self.configure(fg_color=COLORS["gris_bg"])
        self.transient(parent)
        self.grab_set()
        self.after(60, self._maximize)
        self._build()

    def _maximize(self):
        try:
            self.state('zoomed')
        except Exception:
            try:
                self.attributes('-zoomed', True)
            except Exception:
                pass

    def _build(self):
        head = ctk.CTkFrame(self, fg_color=COLORS["orange"], corner_radius=0, height=60)
        head.pack(fill="x")
        head.pack_propagate(False)
        ctk.CTkLabel(head, text="⚡   Nouvelle déclaration de panne / maintenance — ORC1",
                     font=("Segoe UI", 16, "bold"),
                     text_color=COLORS["blanc"], anchor="w"
                     ).pack(side="left", padx=24)

        scroll = ctk.CTkScrollableFrame(self, fg_color=COLORS["gris_bg"], corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=20, pady=12)

        _form_section(scroll, "1", "📅", "Contexte", color=COLORS["orange"])
        s1 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s1.pack(fill="x", pady=(0, 10))
        r1 = ctk.CTkFrame(s1, fg_color="transparent")
        r1.pack(fill="x", padx=16, pady=10)
        self.f_date         = _form_field(r1, "Date *",  width=160, default=date.today().strftime("%d/%m/%Y"))
        self.f_poste        = _form_dropdown(r1, "Poste", LISTES["postes"], width=160, default="Matin")
        self.f_pilote_panne = _form_dropdown(r1, "Pilote de poste", LISTES["pilotes"], width=220)

        _form_section(scroll, "2", "⚡", "Détails de l'intervention", color=COLORS["orange"])
        s2 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s2.pack(fill="x", pady=(0, 10))

        r2a = ctk.CTkFrame(s2, fg_color="transparent")
        r2a.pack(fill="x", padx=16, pady=(10, 0))
        self.f_hdebut     = _form_field(r2a, "Heure début (hh:mm)", width=170)
        self.f_hfin       = _form_field(r2a, "Heure fin (hh:mm)",   width=170)
        self.f_duree_auto = _form_auto(r2a, "Durée calculée", "—", width=140)
        self.f_type       = _form_dropdown(r2a, "Type d'arrêt *",
                                           LISTES["types_panne"], width=200, default="Panne")

        self.f_hdebut.bind("<KeyRelease>", self._recompute)
        self.f_hfin.bind("<KeyRelease>",   self._recompute)

        r2b = ctk.CTkFrame(s2, fg_color="transparent")
        r2b.pack(fill="x", padx=16, pady=(0, 10))
        self.f_equip       = _form_dropdown(r2b, "Équipement concerné *",
                                            LISTES["equipements"], width=280)
        self.f_intervenant = _form_dropdown(r2b, "Intervenant",
                                            LISTES["intervenants"], width=250)

        r2c = ctk.CTkFrame(s2, fg_color="transparent")
        r2c.pack(fill="x", padx=16, pady=(0, 10))
        self.f_detail = _form_field(r2c, "Nature / détail de l'intervention",
                                    width=600, expand=True)

        _form_section(scroll, "3", "💬", "Commentaire général", color=COLORS["gris_label"], optional=True)
        s3 = ctk.CTkFrame(scroll, fg_color=COLORS["blanc"],
                          corner_radius=10, border_width=1, border_color=COLORS["gris_bord"])
        s3.pack(fill="x", pady=(0, 10))
        self.f_comment = ctk.CTkTextbox(s3, height=80,
                                        font=("Segoe UI", 12),
                                        fg_color=COLORS["blanc"],
                                        border_color=COLORS["gris_bord"],
                                        border_width=1, corner_radius=6)
        self.f_comment.pack(fill="x", padx=16, pady=10)

        footer = ctk.CTkFrame(self, fg_color=COLORS["blanc"],
                              corner_radius=0, height=68,
                              border_width=1, border_color=COLORS["gris_bord"])
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        ctk.CTkLabel(footer, text="  * Champs obligatoires",
                     font=("Segoe UI", 11),
                     text_color=COLORS["gris_label"]).pack(side="left", padx=20)

        ctk.CTkButton(footer, text="Annuler",
                      font=("Segoe UI", 12),
                      fg_color=COLORS["gris_bg"], hover_color=COLORS["gris_bord"],
                      text_color=COLORS["gris_label"],
                      border_width=1, border_color=COLORS["gris_bord"],
                      corner_radius=8, height=42, width=110,
                      command=self.destroy).pack(side="right", padx=(0, 16), pady=13)

        ctk.CTkButton(footer, text="✓   Enregistrer la panne",
                      font=("Segoe UI", 13, "bold"),
                      fg_color=COLORS["orange"], hover_color="#C2410C",
                      text_color=COLORS["blanc"],
                      corner_radius=8, height=42, width=220,
                      command=self._save).pack(side="right", padx=(0, 8), pady=13)

    def _recompute(self, *_):
        try:
            hd_str = self.f_hdebut.get().strip().replace("h", ":")
            hf_str = self.f_hfin.get().strip().replace("h", ":")
            if ":" in hd_str and ":" in hf_str:
                pd = hd_str.split(":")
                pf = hf_str.split(":")
                duree = (int(pf[0])*60 + int(pf[1])) - (int(pd[0])*60 + int(pd[1]))
                if duree < 0:
                    duree += 1440
                self.f_duree_auto.configure(text=f"{duree} min  · auto")
                self.f_duree_auto._auto_value = duree
            else:
                self.f_duree_auto.configure(text="—  · auto")
                self.f_duree_auto._auto_value = 0
        except Exception:
            pass

    def _save(self):
        equip = self.f_equip.get()
        if equip == "Sélectionner…" or not equip:
            messagebox.showerror("Champ obligatoire", "L'équipement concerné est obligatoire.")
            return

        date_str = self.f_date.get().strip()
        try:
            d_obj = datetime.strptime(date_str, "%d/%m/%Y").date()
        except ValueError:
            messagebox.showerror("Date invalide", "Format attendu : jj/mm/aaaa")
            return

        duree = getattr(self.f_duree_auto, '_auto_value', 0) or 0

        decl = {h: "" for h in EXCEL_HEADERS}
        decl["Type"]               = "panne"
        decl["Date"]               = d_obj
        decl["Heure"]              = datetime.now().time().replace(microsecond=0)
        decl["Ligne"]              = "ORC1"
        decl["Poste"]              = self.f_poste.get()
        decl["Pilote"]             = self.f_pilote_panne.get() if self.f_pilote_panne.get() != "Sélectionner…" else ""
        decl["N° OF"]              = "—"
        decl["Durée pannes (min)"] = duree
        decl["Nb pannes"]          = 1
        decl["Équipement"]         = equip
        decl["Type panne"]         = self.f_type.get()
        decl["Intervenant"]        = self.f_intervenant.get() if self.f_intervenant.get() != "Sélectionner…" else ""
        decl["Détail panne"]       = self.f_detail.get()
        decl["Commentaire"]        = self.f_comment.get("1.0", "end").strip()

        for val_str, key in [(self.f_hdebut.get(), "Heure début"), (self.f_hfin.get(), "Heure fin")]:
            try:
                p = val_str.replace("h", ":").split(":")
                if len(p) >= 2:
                    decl[key] = time(int(p[0]), int(p[1]))
            except Exception:
                pass

        if self.store.add(decl):
            messagebox.showinfo("Enregistré", "Déclaration de panne enregistrée.")
            if self.on_save:
                self.on_save()
            self.destroy()


# ===========================================================================
# POINT D'ENTRÉE
# ===========================================================================

if __name__ == "__main__":
    app = App()
    app.mainloop()
